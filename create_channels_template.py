"""
create_channels_template.py
~~~~~~~~~~~~~~~~~~~~~~~~~~~
Run once to generate a sample channels.xlsx with the expected columns.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Channels"

# Header row
headers = ["channel_id", "channel_handle", "name"]
header_fill = PatternFill("solid", fgColor="4F81BD")
header_font = Font(bold=True, color="FFFFFF")

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Sample rows
samples = [
    ("", "@MrBeast", "MrBeast"),
    ("UCVHFbw7woebKtFFqtypAKfw", "", "Kurzgesagt"),
    ("", "@veritasium", "Veritasium"),
]
for row_idx, (ch_id, ch_handle, name) in enumerate(samples, 2):
    ws.cell(row=row_idx, column=1, value=ch_id)
    ws.cell(row=row_idx, column=2, value=ch_handle)
    ws.cell(row=row_idx, column=3, value=name)

# Column widths
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 20

output = "channels.xlsx"
wb.save(output)
print(f"Template saved → {output}")
print("Edit the file: add your channel IDs/handles, then run main.py")
